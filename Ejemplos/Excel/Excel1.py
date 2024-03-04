import re
import time
import random
import pytest
import sys
import openpyxl
from playwright.sync_api import Page, expect,Playwright, sync_playwright
from configtest import set_up_excel
from Funciones import Funciones_Globales

# primero correr de formar normal y despues ejecutar el comando(playwright show-trace trace.zip)
# playwright show-trace trace.zip
#pytest Multiples_test.py -s -v
#grabar playwright codegen https://demoqa.com/automation-practice-form

pdf1="C:/Users/rodri/Videos/Playwrite/Curso/Estudiantes/Ejemplos/PDF/dgmn-certificado-de-situacion-militar-17276128.pdf"
rutaexcel="C:/Users/rodri/Videos/Playwrite/Curso/Estudiantes/Ejemplos/Excel/Libro1.xlsx"
registros=3
# Leyendo el archivo excel
archivo=openpyxl.load_workbook(rutaexcel)
def Numero_filas(hoja):
 ac=archivo[hoja]
 return ac.max_row


def Dato_columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

print(Numero_filas("Hoja1")) 
print(Dato_columna("Hoja1",2,2))

def test_excel1(set_up_excel)-> None:
    page=set_up_excel
    F=Funciones_Globales(page)
    F.Scroll_xy(0,600)
    print("cargando")
    F.Esperar(3)
    
    for n in range(2,registros):
        nombre= Dato_columna("Hoja1",n,1)
        ap= Dato_columna("Hoja1",n,2)
        email= Dato_columna("Hoja1",n,3)
    print(nombre+" "+ap+" "+email+" ")
    
    F.Texto("//input[contains(@id,'firstName')]",nombre)
    F.Texto("//input[contains(@id,'lastName')]",ap)
    F.Texto("//input[contains(@id,'userEmail')]",email)
    F.Esperar(3)
    page.goto("https://demoqa.com/automation-practice-form")
    F.Esperar(1)
    





